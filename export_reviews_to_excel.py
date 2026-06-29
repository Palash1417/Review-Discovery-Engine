"""
Export all collected Spotify reviews into a single Excel workbook.

Sheets:
  - All Reviews        : every collected review (raw, deduped)
  - Discovery Reviews  : filtered to music-discovery content
  - Tagged Reviews     : discovery reviews + AI tags (theme/sentiment/pain point)
  - Summary            : counts via live Excel formulas

Output: Spotify_Reviews.xlsx
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

OUT = "Spotify_Reviews.xlsx"
# Upgraded to the full 2,033-review dataset (19,943 raw collected).
SOURCES = [
    ("All Reviews", "spotify_reviews_raw_large.csv"),
    ("Discovery Reviews", "spotify_discovery_reviews_2000.csv"),
    ("Tagged Reviews", "spotify_discovery_reviews_2000_tagged.csv"),
]

HEADER_FILL = PatternFill("solid", start_color="1DB954")  # Spotify green
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
# sensible column widths by header name
WIDTHS = {"source": 14, "date": 22, "text": 90, "rating": 8, "url": 40,
          "discovery_subtheme": 24, "sentiment": 12, "pain_point": 50}


def write_data_sheets() -> dict[str, int]:
    counts = {}
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        for sheet, path in SOURCES:
            if not os.path.exists(path):
                print(f"  skip (missing): {path}")
                continue
            df = pd.read_csv(path)
            df.to_excel(xw, sheet_name=sheet, index=False)
            counts[sheet] = len(df)
            print(f"  {sheet}: {len(df)} rows")
    return counts


def format_workbook() -> None:
    wb = load_workbook(OUT)
    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        for ci, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(str(name), 18)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        # body font (skip huge files cheaply: only set for <= 5000 rows)
        if ws.max_row <= 5000:
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = BODY_FONT
                    if cell.column_letter in ("C", "H"):  # text / pain_point
                        cell.alignment = Alignment(wrap_text=False, vertical="top")
    wb.save(OUT)


def add_summary(counts: dict[str, int]) -> None:
    wb = load_workbook(OUT)
    ws = wb.create_sheet("Summary", 0)  # first sheet
    ws.sheet_view.showGridLines = False

    title = ws["A1"]; title.value = "Spotify Reviews — Summary"
    title.font = Font(name="Arial", bold=True, size=14)

    rows = [
        ("Metric", "Value"),
        ("Total reviews collected", "=COUNTA('All Reviews'!A:A)-1"),
        ("  from Google Play", "=COUNTIF('All Reviews'!A:A,\"play_store\")"),
        ("  from Apple App Store", "=COUNTIF('All Reviews'!A:A,\"app_store\")"),
        ("Discovery-relevant reviews", "=COUNTA('Discovery Reviews'!A:A)-1"),
        ("Tagged reviews", "=COUNTA('Tagged Reviews'!A:A)-1"),
        ("", ""),
        ("Sentiment (tagged)", ""),
        ("  negative", "=COUNTIF('Tagged Reviews'!G:G,\"negative\")"),
        ("  positive", "=COUNTIF('Tagged Reviews'!G:G,\"positive\")"),
        ("  mixed", "=COUNTIF('Tagged Reviews'!G:G,\"mixed\")"),
        ("", ""),
        ("Discovery sub-themes (tagged)", ""),
    ]
    subthemes = ["recommendation_quality", "repetition_filter_bubble",
                 "discovery_features", "playlist_curation", "search_and_navigation",
                 "algorithm_personalization", "catalog_availability", "other"]
    for st in subthemes:
        rows.append((f"  {st}", f"=COUNTIF('Tagged Reviews'!F:F,\"{st}\")"))

    start = 3
    for i, (label, val) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=label).font = (
            Font(name="Arial", bold=True) if (val == "" or label in
            ("Metric",)) else Font(name="Arial"))
        c = ws.cell(row=r, column=2, value=val)
        c.font = Font(name="Arial", bold=(label == "Metric"))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    # Force Excel to recalculate every formula when the file is opened, so the
    # Summary shows values immediately (openpyxl writes formulas without values).
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    wb.save(OUT)


def main() -> int:
    print("Building", OUT)
    counts = write_data_sheets()
    format_workbook()
    add_summary(counts)
    print("Done:", OUT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
